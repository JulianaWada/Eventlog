import pathlib
import tkinter as tk
from time import strftime
from tkinter import *
from tkinter import filedialog, Frame
from tkinter import ttk, RIDGE, CENTER, BOTTOM, VERTICAL, HORIZONTAL, RIGHT, BOTH, Y, X
from tkinter.messagebox import showinfo
import openpyxl
import pandas as pd
import self
from openpyxl import Workbook
from tkcalendar import DateEntry
from openpyxl import Workbook

#===========================================open Downtime===========================

def openDowntime():
    new_window=Toplevel(root)
    new_window.geometry("1100x600+220+130")
    new_window.config(bg="powderblue")
    new_window.focus_force()

    # -------------connecting to the xlsx file-------------

    file = pathlib.Path("Event_log1.xlsx")
    if file.exists():
        pass
    else:
        file = Workbook()
        sheet = file.active
        sheet["A1"] = "Event_ID"
        sheet["B1"] = "Date"
        sheet["C1"] = "Week"
        sheet["D1"] = "Month"
        sheet["E1"] = "Incident"
        sheet["F1"] = "Cluster"
        sheet["G1"] = "Downtime(Start)"
        sheet["H1"] = "Downtime(End)"
        sheet["I1"] = "Total Downtime"
        sheet["L1"] = "SLA%"

        file.save("Event_log1.xlsx")

    def add():
        a = event_id.get()
        b = txt_date.get()
        c = cmd_week.get()
        d = cmd_month.get()
        e = txt_Incident.get()
        f = cmd_cluster.get()
        g = txt_downtime_start.get()
        h = txt_downtime_end.get()
        i = txt_total_downtime.get()

        print(a)
        print(b)
        print(c)
        print(d)
        print(e)
        print(f)
        print(g)
        print(h)
        print(i)

        wb = openpyxl.load_workbook("./Event_log1.xlsx")
        sheet1 = wb.active
        sheet1.cell(column=1, row=sheet1.max_row + 1, value=a)
        sheet1.cell(column=2, row=sheet1.max_row, value=b)
        sheet1.cell(column=3, row=sheet1.max_row, value=c)
        sheet1.cell(column=4, row=sheet1.max_row, value=d)
        sheet1.cell(column=5, row=sheet1.max_row, value=e)
        sheet1.cell(column=6, row=sheet1.max_row, value=f)
        sheet1.cell(column=7, row=sheet1.max_row, value=g)
        sheet1.cell(column=8, row=sheet1.max_row, value=h)
        sheet1.cell(column=9, row=sheet1.max_row, value=i)

        wb.save("Event_log1.xlsx")
        showinfo("saved", "Entry has been saved")

        # xfile = pd.read_excel('Event_log.xlsx', 'sheet')
        # xfile.to_csv('Event_log.csv', index=False)

    def start():
        fp = pd.read_excel("./Event_log1.xlsx")  # Read xlsx file
        for _ in range(len(fp.index.values)):  # use for loop to get values in each line, _ is the number of line.
            EventTable.insert('', 'end', value=tuple(fp.iloc[_, [0, 1, 2, 3, 4, 5, 6, 7, 8,
                                                                 9]].values))  # [_,[1,2]] represents that you will get the values of second column and third column for each line.

        # =====searchFrame=====

    searchFrame = tk.LabelFrame(new_window, text="Search Downtime Event", font=("goudy old style", 10, "bold"), bd=2,
                                relief=RIDGE, bg="white")
    searchFrame.place(x=750, y=20, width=320, height=70)

    # ====options======
    cmd_search = ttk.Combobox(searchFrame, values=("Select", "Event_ID", "Date", "Week", "Month", "Cluster"),
                              state='readonly', justify=CENTER, font=("goudy old style", 10))
    cmd_search.place(x=10, y=10, width=80)
    cmd_search.current(0)

    txt_search = tk.Entry(searchFrame, font=("goudy old style", 10), bg="light yellow").place(x=100, y=10)
    # btn_search = Button(searchFrame,text="Search",command=search,font=("goudy old style", 15), bg="#4caf50",fg="white",cursor="hand2").place(x=410, y=9,width=150,height=30)

    # ======title========
    title = tk.Label(new_window, text="Enter Downtime Event Details", font=("goudy old style", 15), bg="#0f4d7d",
                     fg="powderblue").place(x=10, y=20, width=300)

    # ======content=======
    # ======row1==========
    lbl_evid = tk.Label(new_window, text="Event_ID", font=("goudy old style", 10), bg="powderblue").place(x=50, y=70)
    event_id = tk.Entry(new_window, font=("goudy old style", 10), bg="light yellow")
    event_id.place(x=150, y=70, width=100)

    # =======row2==========
    lbl_date = tk.Label(new_window, text="Date", font=("goudy old style", 10), bg="powderblue").place(x=50, y=110)
    txt_date = DateEntry(new_window, selectmode="day", bg="light yellow", font=("goudy old style", 10))
    txt_date.grid(row=1, column=1, padx=150, pady=110)

    # =======row3==========
    lbl_week = tk.Label(new_window, text="Week", font=("goudy old style", 10), bg="powderblue").place(x=50, y=150)
    cmd_week = ttk.Combobox(new_window, values=("Select", "week1", "week2", "week3", "week4"), state='readonly',
                            justify=CENTER,
                            font=("goudy old style", 10))
    cmd_week.place(x=150, y=150, width=100)
    cmd_week.current(0)

    # =======row4==========
    lbl_month = tk.Label(new_window, text="Month", font=("goudy old style", 10), bg="powderblue").place(x=50, y=190)
    cmd_month = ttk.Combobox(new_window, values=(
    "Select", "January", "February", "April", "May", "June", "July", "August", "September", "October", "November",
    "December"), state='readonly', justify=CENTER, font=("goudy old style", 10))
    cmd_month.place(x=150, y=190, width=100)
    cmd_month.current(0)

    # =======row5==========
    lbl_Incident = tk.Label(new_window, text="Incident", font=("goudy old style", 10), bg="powderblue").place(x=50,
                                                                                                              y=230)
    txt_Incident = tk.Entry(new_window, font=("goudy old style", 10), bg="light yellow")
    txt_Incident.place(x=150, y=230, width=100)

    # =======row6==========
    lbl_cluster = tk.Label(new_window, text="Cluster", font=("goudy old style", 10), bg="powderblue").place(x=50, y=270)
    cmd_cluster = ttk.Combobox(new_window, values=("Select", "OPUS", "GDC", "Shanghai", "FLEX", "costa Rica", "Purley"),
                               state='readonly', justify=CENTER, font=("goudy old style", 10))
    cmd_cluster.place(x=150, y=270, width=100)
    cmd_cluster.current(0)

    # =======row7&8&9==========
    lbl_downtime_start = tk.Label(new_window, text="Downtime(Start)", font=("goudy old style", 10),
                                  bg="powderblue").place(x=50,
                                                         y=310)
    lbl_downtime_end = tk.Label(new_window, text="Downtime(End)", font=("goudy old style", 10), bg="powderblue").place(
        x=50,
        y=350)
    lbl_total_downtime = tk.Label(new_window, text="Total Downtime", font=("goudy old style", 10),
                                  bg="powderblue").place(x=50,
                                                         y=390)

    txt_downtime_start = tk.Entry(new_window, font=("goudy old style", 10), bg="light yellow")
    txt_downtime_start.place(x=150, y=310, width=100)
    txt_downtime_end = tk.Entry(new_window, font=("goudy old style", 10), bg="light yellow")
    txt_downtime_end.place(x=150, y=350, width=100)
    txt_total_downtime = tk.Entry(new_window, font=("goudy old style", 10), bg="light yellow")
    txt_total_downtime.place(x=150, y=390, width=100)

    # ========buttons====
    btn_add = tk.Button(new_window, text="Save", command=add, font=("goudy old style", 10), bg="#4caf50", fg="white",
                        cursor="hand2")
    btn_add.place(x=50, y=440, width=80, height=28)

    btn_open = tk.Button(new_window, text="viewReport", command=start, font=("goudy old style", 10), bg="#009688",
                         fg="white", cursor="hand2")
    btn_open.place(x=150, y=440, width=100, height=28)
    btn_close = tk.Button(new_window, text="close", command=lambda: new_window.destroy(),
                          font=("goudy old style", 10), bg="orange",
                          fg="white", cursor="hand2")
    btn_close.place(x=150, y=480, width=100, height=28)

    # -----------create treeview---------------------
    emp_frame = Frame(new_window, bd=3, relief=RIDGE)
    emp_frame.place(x=290, y=100, width=800, height=450)

    scrolly = tk.Scrollbar(emp_frame, orient=VERTICAL)
    scrollx = tk.Scrollbar(emp_frame, orient=HORIZONTAL)

    EventTable = ttk.Treeview(emp_frame, columns=(
    "Event_ID", "Date", "Week", "Month", "Incident", "Cluster", "Downtime(Start)", "Downtime(End)", "Total Downtime"),
                              yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
    scrollx.pack(side=BOTTOM, fill=X)
    scrolly.pack(side=RIGHT, fill=Y)
    scrollx.config(command=EventTable.xview)
    scrolly.config(command=EventTable.yview)

    EventTable.heading("Event_ID", text="Event_ID")
    EventTable.heading("Date", text="Date")
    EventTable.heading("Week", text="Week")
    EventTable.heading("Month", text="Month")
    EventTable.heading("Incident", text="Incident")
    EventTable.heading("Cluster", text="Cluster")
    EventTable.heading("Downtime(Start)", text="Downtime(Start)")
    EventTable.heading("Downtime(End)", text="Downtime(End)")
    EventTable.heading("Total Downtime", text="Total Downtime")

    EventTable["show"] = "headings"

    EventTable.column("Event_ID", width=20)
    EventTable.column("Date", width=20)
    EventTable.column("Week", width=20)
    EventTable.column("Month", width=20)
    EventTable.column("Incident", width=20)
    EventTable.column("Cluster", width=20)
    EventTable.column("Downtime(Start)", width=20)
    EventTable.column("Downtime(End)", width=20)
    EventTable.column("Total Downtime", width=20)
    EventTable.pack(fill=BOTH, expand=1)
    EventTable.bind("<ButtonRelease-1>")

#==============================================open SLA===============================================

def openSla():
    new_window1 = Toplevel(root)
    new_window1.geometry("1100x600+220+130")
    new_window1.config(bg="powderblue")
    new_window1.focus_force()

    fl = pathlib.Path("Event_log1.xlsx")
    if fl.exists():
        pass
    else:
        fl= Workbook()
        sheet2 = fl.active
        sheet2["A1"] = "Event_ID"
        sheet2["B1"] = "Date"
        sheet2["C1"] = "Week"
        sheet2["D1"] = "Month"
        sheet2["E1"] = "Incident"
        sheet2["F1"] = "Cluster"
        sheet2["G1"] = "Downtime(Start)"
        sheet2["H1"] = "Downtime(End)"
        sheet2["I1"] = "Total Downtime"


        file.save("Event_log1.xlsx")

    def start():
        fp = pd.read_excel("./Event_log1.xlsx")  # Read xlsx file
        for _ in range(len(fp.index.values)):  # use for loop to get values in each line, _ is the number of line.
            DailyTable.insert('', 'end', value=tuple(fp.iloc[_, [15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,
                                                                 27]].values))  # [_,[1,2]] represents that you will get the values of second column and third column for each line.

    # =======SLA FRAME Functions====================

    def submit():
        sb = pd.read_excel("./Event_log1.xlsx")  # Read xlsx file
        for _ in range(len(sb.index.values)):  # use for loop to get values in each line, _ is the number of line.
            slaTable.insert('', 'end', value=tuple(sb.iloc[_, [10, 11, 12,
                                                               13]].values))  # [_,[1,2]] represents that you will get the values of second column and third column for each line.

        # =====searchFrame=====

    searchFrame = tk.LabelFrame(new_window1, text="Search Daily SLA's Event", font=("goudy old style", 10, "bold"),
                                bd=2, relief=RIDGE, bg="white")
    searchFrame.place(x=750, y=20, width=320, height=70)

    # ====options======
    cmd_search = ttk.Combobox(searchFrame, values=("Select", "Date"), state='readonly', justify=CENTER,
                              font=("goudy old style", 10))
    cmd_search.place(x=10, y=10, width=80)
    cmd_search.current(0)

    txt_search = tk.Entry(searchFrame, font=("goudy old style", 10), bg="light yellow").place(x=100, y=10)
    btn_search = Button(searchFrame,text="Search",font=("goudy old style", 15), bg="#4caf50",fg="white",cursor="hand2").place(x=410, y=9,width=150,height=30)

    # ==============SLA report Frame==========
    slaFrame = tk.LabelFrame(new_window1, text="Clusters Monthly SLA in %", font=("goudy old style", 9, "bold"), bd=2,
                             bg="light blue", relief=RIDGE)
    slaFrame.place(x=10, y=100, width=290, height=350)

    lbl_mon = tk.Label(new_window1, text="Months:", font=("goudy old style", 9), bg="green", fg="powderblue").place(
        x=50, y=80, width=50)
    month = DateEntry(new_window1, selectmode="day", year=2022, month=1, day=1)
    month.grid(row=1, column=1, padx=120, pady=80)
    scrolly = tk.Scrollbar(slaFrame, orient=VERTICAL)
    scrollx = tk.Scrollbar(slaFrame, orient=HORIZONTAL)

    slaTable = ttk.Treeview(slaFrame, columns=("MONTH", "CLUSTER", "SLA%", "DOWNTIME"), yscrollcommand=scrolly.set,
                            xscrollcommand=scrollx.set)
    scrollx.pack(side=BOTTOM, fill=X)
    scrolly.pack(side=RIGHT, fill=Y)
    scrollx.config(command=slaTable.xview)
    scrolly.config(command=slaTable.yview)

    slaTable.heading("MONTH", text="MONTH")
    slaTable.heading("CLUSTER", text="CLUSTER")
    slaTable.heading("SLA%", text="SLA%")
    slaTable.heading("DOWNTIME", text="DOWNTIME")
    slaTable["show"] = "headings"

    slaTable.column("MONTH", width=20)
    slaTable.column("CLUSTER", width=20)
    slaTable.column("SLA%", width=20)
    slaTable.column("DOWNTIME", width=20)
    slaTable.pack(fill=BOTH, expand=1)
    slaTable.bind("<ButtonRelease-1>")

    # ===================BUTTONS==================
    btn_open = tk.Button(new_window1, text="Show Monthly", command=submit, font=("goudy old style", 10), bg="green",
                         fg="white",
                         cursor="hand2")
    btn_open.place(x=120, y=460, width=100, height=28)

    btn_open = tk.Button(new_window1, text="Show Daily", command=start, font=("goudy old style", 10), bg="#009688",
                         fg="white", cursor="hand2")
    btn_open.place(x=500, y=460, width=100, height=28)
    btn_close = tk.Button(new_window1, text="close", command=lambda: new_window1.destroy(),
                          font=("goudy old style", 10), bg="orange",
                          fg="white", cursor="hand2")
    btn_close.place(x=500, y=500, width=100, height=28)

    # -----------create treeview---------------------
    emp_frame = tk.LabelFrame(new_window1, text="Clusters SLA's in %", font=("goudy old style", 9, "bold"), bd=3,
                              bg="light blue", relief=RIDGE)
    emp_frame.place(x=320, y=100, width=780, height=350)

    scrolly = tk.Scrollbar(emp_frame, orient=VERTICAL)
    scrollx = tk.Scrollbar(emp_frame, orient=HORIZONTAL)

    DailyTable = ttk.Treeview(emp_frame, columns=(
    "MONTH", "OPUS", "OPUS_D", "GDC", "GDC_D", "SHANGHAI", "SHANGHAI_D", "FLEX", "FLEX_D", "COSTA RICA", "COSTA RICA_D",
    "PURLEY", "PURLEY_D"),
                              yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
    scrollx.pack(side=BOTTOM, fill=X)
    scrolly.pack(side=RIGHT, fill=Y)
    scrollx.config(command=DailyTable.xview)
    scrolly.config(command=DailyTable.yview)

    DailyTable.heading("MONTH", text="DATE")
    DailyTable.heading("OPUS", text="OPUS")
    DailyTable.heading("OPUS_D", text="OPUS_D")
    DailyTable.heading("GDC", text="GDC")
    DailyTable.heading("GDC_D", text="GDC_D")
    DailyTable.heading("SHANGHAI", text="SHANGHAI")
    DailyTable.heading("SHANGHAI_D", text="SHANGHAI_D")
    DailyTable.heading("FLEX", text="FLEX")
    DailyTable.heading("FLEX_D", text="FLEX_D")
    DailyTable.heading("COSTA RICA_D", text="COSTA RICA_D")
    DailyTable.heading("COSTA RICA_D", text="COSTA RICA_D")
    DailyTable.heading("PURLEY", text="PURLEY")
    DailyTable.heading("PURLEY_D", text="PURLEY_D")

    DailyTable["show"] = "headings"
    DailyTable.column("MONTH", width=100)
    DailyTable.column("OPUS", width=100)
    DailyTable.column("OPUS_D", width=100)
    DailyTable.column("GDC", width=100)
    DailyTable.column("GDC_D", width=100)
    DailyTable.column("SHANGHAI", width=100)
    DailyTable.column("SHANGHAI_D", width=100)
    DailyTable.column("FLEX", width=100)
    DailyTable.column("FLEX_D", width=100)
    DailyTable.column("COSTA RICA", width=100)
    DailyTable.column("COSTA RICA_D", width=100)
    DailyTable.column("PURLEY", width=100)
    DailyTable.column("PURLEY_D", width=100)
    DailyTable.pack(fill=BOTH, expand=1)
    DailyTable.bind("<ButtonRelease-1>")

#=============================================open MTBF and MTTR===============================================
def openMean():
    new_window2 = Toplevel(root)
    new_window2.geometry("1100x600+220+130")
    new_window2.config(bg="powderblue")
    new_window2.focus_force()

    file = pathlib.Path("MTBF.xlsx")
    if file.exists():
        pass
    else:
        wb1 = Workbook()
        sheet3 = wb1.active
        sheet3["A1"] = "Event_ID"
        sheet3["B1"] = "Date"
        sheet3["C1"] = "Week"
        sheet3["D1"] = "Month"
        sheet3["E1"] = "Incident"
        sheet3["F1"] = "Cluster"
        sheet3["G1"] = "Repair time(start)"
        sheet3["H1"] = "Repair time(end)"
        sheet3["I1"] = "Time since last failure"
        sheet3["J1"] = "Time to repair"

        wb1.save("MTBF.xlsx")

    def add():
        a = event_id.get()
        b = txt_date.get()
        c = cmd_week.get()
        d = cmd_month.get()
        e = txt_Incident.get()
        f = cmd_cluster.get()
        g = txt_startrepair_time.get()
        h = txt_endrepair_time.get()
        i = txt_Time_sincefailure.get()
        j = txt_Time_torepair.get()

        print(a)
        print(b)
        print(c)
        print(d)
        print(e)
        print(f)
        print(g)
        print(h)
        print(i)
        print(j)

        wb3 = openpyxl.load_workbook("MTBF.xlsx")
        sheet4 = wb3.active
        sheet4.cell(column=1, row=sheet.max_row + 1, value=a)
        sheet4.cell(column=2, row=sheet.max_row, value=b)
        sheet4.cell(column=3, row=sheet.max_row, value=c)
        sheet4.cell(column=4, row=sheet.max_row, value=d)
        sheet4.cell(column=5, row=sheet.max_row, value=e)
        sheet4.cell(column=6, row=sheet.max_row, value=f)
        sheet4.cell(column=7, row=sheet.max_row, value=g)
        sheet4.cell(column=8, row=sheet.max_row, value=h)
        sheet4.cell(column=8, row=sheet.max_row, value=i)
        sheet4.cell(column=8, row=sheet.max_row, value=j)

        wb3.save("MTBF.xlsx")
        showinfo("saved", "Entry has been saved")

        # xfile = pd.read_excel('Event_log.xlsx', 'sheet')
        # xfile.to_csv('Event_log.csv', index=False)

    def start():
        fp = pd.read_excel("./MTBF.xlsx")  # Read xlsx file
        for _ in range(len(fp.index.values)):  # use for loop to get values in each line, _ is the number of line.
            mmTable.insert('', 'end', value=tuple(fp.iloc[_, [0, 1, 2, 3, 4, 5, 6, 7, 8,
                                                              9]].values))  # [_,[1,2]] represents that you will get the values of second column and third column for each line.

    def submit():
        sb = pd.read_excel("./MTBF.xlsx")  # Read xlsx file
        for _ in range(len(sb.index.values)):  # use for loop to get values in each line, _ is the number of line.
            meanTable.insert('', 'end', value=tuple(sb.iloc[_, [11, 12, 13]].values))

        # =====searchFrame=====

    searchFrame = tk.LabelFrame(new_window2, text="Search Repair Event", font=("goudy old style", 10, "bold"), bd=2,
                                relief=RIDGE, bg="white")
    searchFrame.place(x=750, y=20, width=320, height=70)

    # ====options======
    cmd_search = ttk.Combobox(searchFrame, values=("Select", "Event_ID", "Date", "Week", "Month", "Cluster"),
                              state='readonly', justify=CENTER, font=("goudy old style", 10))
    cmd_search.place(x=10, y=10, width=80)
    cmd_search.current(0)

    txt_search = tk.Entry(searchFrame, font=("goudy old style", 10), bg="light yellow").place(x=100, y=10)
    # btn_search = Button(searchFrame,text="Search",command=search,font=("goudy old style", 15), bg="#4caf50",fg="white",cursor="hand2").place(x=410, y=9,width=150,height=30)

    # ======title========
    title = tk.Label(new_window2, text="Enter Repair Event Details", font=("goudy old style", 15), bg="#0f4d7d",
                     fg="powderblue").place(x=10, y=20, width=300)

    # ======content=======
    # ======row1==========
    lbl_evid = tk.Label(new_window2, text="Event_ID", font=("goudy old style", 10), bg="powderblue").place(x=50, y=70)
    event_id = tk.Entry(new_window2, font=("goudy old style", 10), bg="light yellow")
    event_id.place(x=150, y=70, width=100)

    # =======row2==========
    lbl_date = tk.Label(new_window2, text="Date", font=("goudy old style", 10), bg="powderblue").place(x=50, y=110)
    txt_date = DateEntry(new_window2, selectmode="day", bg="light yellow", font=("goudy old style", 10))
    txt_date.grid(row=1, column=1, padx=150, pady=110)

    # =======row3==========
    lbl_week = tk.Label(new_window2, text="Week", font=("goudy old style", 10), bg="powderblue").place(x=50, y=150)
    cmd_week = ttk.Combobox(new_window2, values=("Select", "week1", "week2", "week3", "week4"), state='readonly',
                            justify=CENTER,
                            font=("goudy old style", 10))
    cmd_week.place(x=150, y=150, width=100)
    cmd_week.current(0)

    # =======row4==========
    lbl_month = tk.Label(new_window2, text="Month", font=("goudy old style", 10), bg="powderblue").place(x=50, y=190)
    cmd_month = ttk.Combobox(new_window2, values=(
    "Select", "January", "February", "April", "May", "June", "July", "August", "September", "October", "November",
    "December"), state='readonly', justify=CENTER, font=("goudy old style", 10))
    cmd_month.place(x=150, y=190, width=100)
    cmd_month.current(0)

    # =======row5==========
    lbl_Incident = tk.Label(new_window2, text="Incident", font=("goudy old style", 10), bg="powderblue").place(x=50,
                                                                                                               y=230)
    txt_Incident = tk.Entry(new_window2, font=("goudy old style", 10), bg="light yellow")
    txt_Incident.place(x=150, y=230, width=100)

    # =======row6==========
    lbl_cluster = tk.Label(new_window2, text="Cluster", font=("goudy old style", 10), bg="powderblue").place(x=50,
                                                                                                             y=270)
    cmd_cluster = ttk.Combobox(new_window2,
                               values=("Select", "OPUS", "GDC", "Shanghai", "FLEX", "costa Rica", "Purley"),
                               state='readonly', justify=CENTER, font=("goudy old style", 12))
    cmd_cluster.place(x=150, y=270, width=100)
    cmd_cluster.current(0)

    # =======row7&8&9==========
    lbl_startrepair_time = tk.Label(new_window2, text="Repair time(start)", font=("goudy old style", 10),
                                    bg="powderblue").place(x=50,
                                                           y=310)
    lbl_endrepair_time = tk.Label(new_window2, text="Repair time(end)", font=("goudy old style", 10),
                                  bg="powderblue").place(x=50,
                                                         y=350)
    lbl_Time_sincefailure = tk.Label(new_window2, text="Time since last failure", font=("goudy old style", 10),
                                     bg="powderblue").place(x=50,
                                                            y=390)
    lbl_Time_torepair = tk.Label(new_window2, text="Time since last failure", font=("goudy old style", 10),
                                 bg="powderblue").place(x=50,
                                                        y=430)

    txt_startrepair_time = tk.Entry(new_window2, font=("goudy old style", 10), bg="light yellow")
    txt_startrepair_time.place(x=180, y=310, width=100)
    txt_endrepair_time = tk.Entry(new_window2, font=("goudy old style", 10), bg="light yellow")
    txt_endrepair_time.place(x=180, y=350, width=100)
    txt_Time_sincefailure = tk.Entry(new_window2, font=("goudy old style", 10), bg="light yellow")
    txt_Time_sincefailure.place(x=180, y=390, width=100)
    txt_Time_torepair = tk.Entry(new_window2, font=("goudy old style", 10), bg="light yellow")
    txt_Time_torepair.place(x=180, y=430, width=100)
    # ========buttons====
    btn_add = tk.Button(new_window2, text="Save", command=add, font=("goudy old style", 10), bg="#009688", fg="white",cursor="hand2")
    btn_add.place(x=50, y=480, width=80, height=28)

    btn_open = tk.Button(new_window2, text="veiwReport", command=start, font=("goudy old style", 10), bg="#009688",fg="white", cursor="hand2")
    btn_open.place(x=150, y=480, width=100, height=28)
    btn_close = tk.Button(new_window2, text="close", command=lambda: new_window2.destroy(), font=("goudy old style", 10), bg="orange",fg="white", cursor="hand2")
    btn_close.place(x=150, y=520, width=100, height=28)

    # ==============SLA report Frame==========
    meanFrame = tk.LabelFrame(new_window2, text="CLUSTERS MTBF & MTTR", font=("goudy old style", 9, "bold"), bd=2,
                              bg="light blue", relief=RIDGE)
    meanFrame.place(x=300, y=100, width=180, height=150)

    # lbl_mon=tk.Label(slaFrame,text="Months:",font=("goudy old style",9),bg="green",fg="powderblue").place(x=10,y=10,width=50)
    # month=DateEntry(slaFrame,selectmode="day",year=2022,month=1,day=1)
    # month.grid(row=1,column=1,padx=80,pady=10)
    scrolly = tk.Scrollbar(meanFrame, orient=VERTICAL)
    scrollx = tk.Scrollbar(meanFrame, orient=HORIZONTAL)

    meanTable = ttk.Treeview(meanFrame, columns=("_MM", "MTBF", "MTTR"), yscrollcommand=scrolly.set,
                             xscrollcommand=scrollx.set)
    scrollx.pack(side=BOTTOM, fill=X)
    scrolly.pack(side=RIGHT, fill=Y)
    scrollx.config(command=meanTable.xview)
    scrolly.config(command=meanTable.yview)

    meanTable.heading("_MM", text="_MM")
    meanTable.heading("MTBF", text="MTBF")
    meanTable.heading("MTTR", text="MTTR")
    meanTable["show"] = "headings"

    meanTable.column("_MM", width=20)
    meanTable.column("MTBF", width=20)
    meanTable.column("MTTR", width=20)
    meanTable.pack(fill=BOTH, expand=1)
    meanTable.bind("<ButtonRelease-1>")

    btn_open = tk.Button(new_window2, text="compute MTBF & MTTR", command=submit, font=("goudy old style", 10),
                         bg="green", fg="white",
                         cursor="hand2")
    btn_open.place(x=300, y=250, width=150, height=28)

    # -----------create treeview---------------------
    emp_frame = Frame(new_window2, bd=3, relief=RIDGE)
    emp_frame.place(x=500, y=100, width=600, height=450)

    scrolly = tk.Scrollbar(emp_frame, orient=VERTICAL)
    scrollx = tk.Scrollbar(emp_frame, orient=HORIZONTAL)

    mmTable = ttk.Treeview(emp_frame, columns=(
    "Event_ID", "Date", "Week", "Month", "Incident", "Cluster", "Repair time(start)", "Repair time(end)",
    "Time since last failure", "Time to repair"),
                           yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
    scrollx.pack(side=BOTTOM, fill=X)
    scrolly.pack(side=RIGHT, fill=Y)
    scrollx.config(command=mmTable.xview)
    scrolly.config(command=mmTable.yview)

    mmTable.heading("Event_ID", text="Event_ID")
    mmTable.heading("Date", text="Date")
    mmTable.heading("Week", text="Week")
    mmTable.heading("Month", text="Month")
    mmTable.heading("Incident", text="Incident")
    mmTable.heading("Cluster", text="Cluster")
    mmTable.heading("Repair time(start)", text="Repair time(start)")
    mmTable.heading("Repair time(end)", text="Repair time(end)")
    mmTable.heading("Time since last failure", text="Time since last failure")
    mmTable.heading("Time to repair", text="Time to repair")

    mmTable["show"] = "headings"

    mmTable.column("Event_ID", width=100)
    mmTable.column("Date", width=100)
    mmTable.column("Week", width=100)
    mmTable.column("Month", width=100)
    mmTable.column("Incident", width=100)
    mmTable.column("Cluster", width=100)
    mmTable.column("Repair time(start)", width=100)
    mmTable.column("Repair time(end)", width=100)
    mmTable.column("Time since last failure", width=100)
    mmTable.column("Time to repair", width=100)
    mmTable.pack(fill=BOTH, expand=1)
    mmTable.bind("<ButtonRelease-1>")

#======================================open Dashboard ============================
root=tk.Tk()
root.geometry("1350x700+0+0")
root.title("Event Metircs | developed by Eng.Samuel")

file = pathlib.Path("Event_log1.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Event_ID"
    sheet["B1"] = "Date"
    sheet["C1"] = "Week"
    sheet["D1"] = "Month"
    sheet["E1"] = "Incident"
    sheet["F1"] = "Cluster"
    sheet["G1"] = "Downtime(Start)"
    sheet["H1"] = "Downtime(End)"
    sheet["I1"] = "Total Downtime"
    sheet["L1"]="SLA%"

    file.save("Event_log1.xlsx")


        #===title=====
title=tk.Label(root,text="Event Log Monitoring system",compound=LEFT,font=("times new roman",30,"bold"),bg="#004aad",fg="white",anchor="w",padx=480).place(x=0,y=0,relwidth=1,height=70)



def my_time():
    time_string=strftime('Dashboard\t\t Date: %x \t\t Time:%H:%M:%S %p')
    lbl_clock.config(text=time_string)
    lbl_clock.after(1000,my_time)
lbl_clock =tk.Label(root,font=("times new roman",15),bg="#4d636d", fg="white", anchor="w", padx=20)
lbl_clock.place(x=0,y=70,relwidth=1, height=30)
my_time()
        #=====left menu===
#MenuLogo=PhotoImage(file="images/m")
#self.MenuLogo=self.MenuLogo.resize((200,200),Image.ANTIALIAS)
#self.MenuLogo= ImageTk.PhotoImage(self.MenuLogo)

LeftMenu=tk.Frame(root,bd=2,relief=RIDGE,bg="white")
LeftMenu.place(x=0,y=102,width=200,height=565)

#lbl_menuLogo=tk.Label(LeftMenu,image=MenuLogo)
#lbl_menuLogo.pack(side=TOP, fill=X)

#icon_side = PhotoImage(file="images/arrow.png")

lbl_menu=tk.Label(LeftMenu,text="menu",font=("times new roman", 20),bg="#009688").pack(side=TOP,fill=X)
btn_employee = tk.Button(LeftMenu, text="DOWNTIME",command=openDowntime,compound=LEFT,padx=5,anchor="w", font=("times new roman", 15,"bold"), bg="white",bd=3,cursor="hand2").pack(side=TOP, fill=X)
btn_supplier= tk.Button(LeftMenu, text="SLA",command=openSla, compound=LEFT, padx=5, anchor="w",font=("times new roman", 15, "bold"), bg="white", bd=3, cursor="hand2").pack(side=TOP,fill=X)
btn_category= tk.Button(LeftMenu, text="MTTR & MTBF",command=openMean, compound=LEFT, padx=5, anchor="w", font=("times new roman", 15, "bold"), bg="white", bd=3, cursor="hand2").pack(side=TOP,fill=X)

        #===content===

opusFrame= tk.LabelFrame(root, text="OPUS DAILY SLA in %", font=("goudy old style", 9, "bold"), bd=2, bg="light blue", relief=RIDGE)
opusFrame.place(x=220, y=120, width=350, height=200)

GDCFrame= tk.LabelFrame(root, text="GDC DAILY SLA in %", font=("goudy old style", 9, "bold"), bd=2, bg="#607d8b", relief=RIDGE)
GDCFrame.place(x=600, y=120, width=350, height=200)

shanghaiFrame= tk.LabelFrame(root, text="Shanghai DAILY SLA in %", font=("goudy old style", 9, "bold"), bd=2, bg="#607d8b", relief=RIDGE)
shanghaiFrame.place(x=980, y=120, width=350, height=200)

flexFrame= tk.LabelFrame(root, text="FLEX DAILY SLA in %", font=("goudy old style", 9, "bold"), bd=2, bg="light blue", relief=RIDGE)
flexFrame.place(x=220, y=350, width=350, height=200)

costaricaFrame= tk.LabelFrame(root, text="COSTA RICA DAILY SLA in %", font=("goudy old style", 9, "bold"), bd=2, bg="#009688", relief=RIDGE)
costaricaFrame.place(x=600, y=350, width=350, height=200)

purleyFrame= tk.LabelFrame(root, text="PURLEY DAILY SLA in %", font=("goudy old style", 9, "bold"), bd=2, bg="#009688", relief=RIDGE)
purleyFrame.place(x=980, y=350, width=350, height=200)





        #===footer===
lbl_footer= Label(root,bg="#4d636d", fg="white").pack(side=BOTTOM,fill=X)
#=======================================================================




root.mainloop()